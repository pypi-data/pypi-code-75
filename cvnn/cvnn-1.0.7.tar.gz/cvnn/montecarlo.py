import logging
import os
import tensorflow as tf
import pandas as pd
import numpy as np
from tqdm import tqdm
from pdb import set_trace
from time import sleep
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
from tensorflow.keras.losses import categorical_crossentropy
from tensorflow.keras import Model
# Own modules
import cvnn
import cvnn.layers as layers
import cvnn.dataset as dp
from cvnn.data_analysis import MonteCarloAnalyzer, Plotter
from cvnn.layers import ComplexDense, ComplexDropout
from cvnn.utils import transform_to_real, randomize
from cvnn.real_equiv_tools import get_real_equivalent
from cvnn.utils import median_error
# typing
from pathlib import Path
from typing import Union, Optional, List, Tuple
from cvnn.activations import t_activation
from tensorflow import data
from typing import Type

logger = logging.getLogger(cvnn.__name__)

t_path = Union[str, Path]


class MonteCarlo:

    def __init__(self):
        """
        Class that allows the statistical comparison of several models on the same dataset
        """
        self.models = []
        self.pandas_full_data = pd.DataFrame()
        self.monte_carlo_analyzer = MonteCarloAnalyzer()  # All at None

    def add_model(self, model: Type[Model]):
        """
        Adds a cvnn.CvnnModel to the list to then comparate between them
        """
        self.models.append(model)

    def run(self, x, y, data_summary: str = '', polar: bool = False, do_conf_mat: bool = True,
            validation_split: float = 0.2,
            validation_data: Optional[Union[Tuple[np.ndarray, np.ndarray], data.Dataset]] = None,    # TODO: Add the tuple of validation data details.
            iterations: int = 100, epochs: int = 10, batch_size: int = 100,
            shuffle: bool = True, debug: bool = False, display_freq: int = 1, checkpoints: bool = False):
        """
        This function is used to compare all models added with `self.add_model` method.
        Runs the iteration dataset (x, y).
        1. It then runs a monte carlo simulation of several iterations of both CVNN and an equivalent RVNN model.
        2. Saves several files into ./log/montecarlo/date/of/run/
            2.1. run_summary.txt: Summary of the run models and data
            2.2. run_data.csv: Full information of performance of iteration of each model at each epoch
            2.3. <model.name>_network_statistical_result.csv: Statistical results of all iterations of CVNN per epoch
            2.4. (Optional) `plot/` folder with the corresponding plots generated by MonteCarloAnalyzer.do_all()
        :param x: Input data. It could be:
            - A Numpy array (or array-like), or a list of arrays (in case the model has multiple inputs).
            - A TensorFlow tensor, or a list of tensors (in case the model has multiple inputs).
            - A tf.data dataset. Should return a tuple (inputs, targets). Preferred data type (less overhead).
        :param y: Labels/Target data. Like the input data x, it could be either Numpy array(s) or TensorFlow tensor(s).
            If f x is a dataset then y will be ignored (default None)
        :param data_summary:  (String) Dataset name to keep track of it
        :param polar: (Boolean) If the model is real.
            Separate the complex data into real and imaginary part (polar = False) or amplitude and phase (polar = True)
        :param do_conf_mat: Generate a confusion matrix based on results.
        :param validation_split: Float between 0 and 1.
            Percentage of the input data to be used as test set (the rest will be use as train set)
            Default: 0.0 (No validation set).
            This input is ignored if validation_data is given.
        :param validation_data: Data on which to evaluate the loss and any model metrics at the end of each epoch.
            The model will not be trained on this data. This parameter takes precedence over validation_split.
            It can be:
                - tuple (x_val, y_val) of Numpy arrays or tensors. Preferred data type (less overhead).
                - A tf.data dataset.
        :param iterations: Number of iterations to be done for each model
        :param epochs: Number of epochs for each iteration
        :param batch_size: Batch size at each iteration
        :param display_freq: Integer (Default 1). Only relevant if validation data is provided.
            Frequency on terms of epochs before running the validation.
        :param shuffle: (Boolean) Whether to shuffle the training data before each epoch.
        :param debug:
        :param checkpoints:
        :return: (string) Full path to the run_data.csv generated file.
            It can be used by cvnn.data_analysis.SeveralMonteCarloComparison to compare several runs.
        """
        x, y = randomize(x, y)
        w_save = []
        for model in self.models:       # ATTENTION: This will omake all models have the SAME weights, not ideal
            w_save.append(model.get_weights())     # Save model weights
        # Reset data frame
        self.pandas_full_data = pd.DataFrame()
        self.save_summary_of_run(self._run_summary(iterations, epochs, batch_size, shuffle),
                                 data_summary)
        if not debug:
            pbar = tqdm(total=iterations)
        for it in range(iterations):
            if debug:
                logger.info("Iteration {}/{}".format(it + 1, iterations))
            if shuffle:  # shuffle all data at each iteration
                x, y = randomize(x, y)
            for i, model in enumerate(self.models):
                val_data_fit = None
                if model.inputs[0].dtype.is_complex:
                    x_fit = x
                    if validation_data is not None:
                        val_data_fit = validation_data
                else:
                    x_fit = transform_to_real(x, polar=polar)
                    if validation_data is not None:
                        val_data_fit = (transform_to_real(validation_data[0], polar=polar), validation_data[1])
                if validation_data is not None:
                    validation_split = 0.0
                model.set_weights(w_save[i])
                run_result = model.fit(x_fit, y, validation_split=validation_split, validation_data=val_data_fit,
                                       epochs=epochs, batch_size=batch_size,
                                       verbose=debug, validation_freq=display_freq)
                # TODO: Must have save_csv_history to do the montecarlo results latter
                # 
                # Save all results
                temp_path = self.monte_carlo_analyzer.path / f"run/iteration{it}_model{i}_{model.name}"
                os.makedirs(temp_path, exist_ok=True)
                plotter = Plotter(path=temp_path, data_results_dict=run_result.history, model_name=model.name)
                self.pandas_full_data = pd.concat([self.pandas_full_data, plotter.get_full_pandas_dataframe()],
                                                  sort=False)
                with open(temp_path / f'{model.name}_metadata.txt', 'w') as fh:
                    model.summary(print_fn=lambda x: fh.write(x + '\n'))
            if not debug:
                pbar.update()
            if checkpoints:
                # Save checkpoint in case Monte Carlo stops in the middle
                self.pandas_full_data.to_csv(self.monte_carlo_analyzer.path / "run_data.csv", index=False)
        if not debug:
            pbar.close()
        self.pandas_full_data = self.pandas_full_data.reset_index(drop=True)
        self.monte_carlo_analyzer.set_df(self.pandas_full_data)

    @staticmethod
    def _run_summary(iterations: int, epochs: int, batch_size: int, shuffle: bool) -> str:
        ret_str = "Monte Carlo run\n"
        ret_str += f"\tIterations: {iterations}\n"
        ret_str += f"\tepochs: {epochs}\n"
        ret_str += f"\tbatch_size: {batch_size}\n"
        if shuffle:
            ret_str += "\tShuffle data at each iteration\n"
        else:
            ret_str += "\tData is not shuffled at each iteration\n"
        return ret_str

    def save_summary_of_run(self, run_summary, data_summary):
        with open(str(self.monte_carlo_analyzer.path / "run_summary.txt"), "w") as file:
            file.write(run_summary)
            file.write(data_summary)
            file.write("Models:\n")
            for model in self.models:
                model.summary(print_fn=lambda x: file.write(x + '\n'))


class RealVsComplex(MonteCarlo):
    """
    Inherits from MonteCarlo. Compares a complex model with it's real equivalent.

    Example usage:
    ```
    # Assume you already have complex data 'x' with its labels 'y'... and a Cvnn model.

    montecarlo = RealVsComplex(complex_model)
    montecarlo.run(x, y)
    ```
    """

    def __init__(self, complex_model: Type[Model], capacity_equivalent: bool = True, equiv_technique: str = 'ratio'):
        """
        :param complex_model: cvnn.CvnnModel
        :param capacity_equivalent: An equivalent model can be equivalent in terms of layer neurons or
                        trainable parameters (capacity equivalent according to: https://arxiv.org/abs/1811.12351)
            - True, it creates a capacity-equivalent model in terms of trainable parameters
            - False, it will double all layer size (except the last one if classifier=True)
        :param equiv_technique: Used to define the strategy of the capacity equivalent model.
            This parameter is ignored if capacity_equivalent=False
            - 'ratio': neurons_real_valued_layer[i] = r * neurons_complex_valued_layer[i], 'r' constant for all 'i'
            - 'alternate': Method described in https://arxiv.org/abs/1811.12351 where one alternates between
                    multiplying by 2 or 1. Special case on the middle is treated as a compromise between the two.
        """
        super().__init__()
        # add models
        self.add_model(complex_model)
        self.add_model(get_real_equivalent(complex_model, capacity_equivalent=capacity_equivalent,
                                           equiv_technique=equiv_technique, name="real_network"))


# ====================================
#   Monte Carlo simulation methods
# ====================================
def run_montecarlo(models: List[Model], dataset: cvnn.dataset.Dataset, open_dataset: Optional[t_path] = None,
                   iterations: int = 500,
                   epochs: int = 150, batch_size: int = 100, display_freq: int = 1,
                   validation_split: float = 0.2,
                   validation_data: Optional[Union[Tuple, data.Dataset]] = None,  # TODO: Add vallidation data tuple details
                   debug: bool = False, polar: bool = False, do_all: bool = True, do_conf_mat: bool = True) -> str:
    """
    This function is used to compare different neural networks performance.
    1. Runs simulation and compares them.
    2. Saves several files into ./log/montecarlo/date/of/run/
        2.1. run_summary.txt: Summary of the run models and data
        2.2. run_data.csv: Full information of performance of iteration of each model at each epoch
        2.3. <model_name>_statistical_result.csv: Statistical results of all iterations of each model per epoch
        2.4. (Optional) `plot/` folder with the corresponding plots generated by MonteCarloAnalyzer.do_all()

    :param models: List of cvnn.CvnnModel to be compared.
    :param dataset: cvnn.dataset.Dataset with the dataset to be used on the training
    :param open_dataset: (None)
        If dataset is saved inside a folder and must be opened, path of the Dataset to be opened. Else None (default)
    :param iterations: Number of iterations to be done for each model
    :param epochs: Number of epochs for each iteration
    :param batch_size: Batch size at each iteration
    :param display_freq: Frequency in terms of epochs of when to do a checkpoint.
    :param debug:
    :param polar: Boolean weather the RVNN should receive real and imaginary part (False) or amplitude and phase (True)
    :param do_all: If true (default) it creates a `plot/` folder with the plots generated by MonteCarloAnalyzer.do_all()
    :param validation_split: Float between 0 and 1.
            Percentage of the input data to be used as test set (the rest will be use as train set)
            Default: 0.0 (No validation set).
            This input is ignored if validation_data is given.
        :param validation_data: Data on which to evaluate the loss and any model metrics at the end of each epoch.
            The model will not be trained on this data. This parameter takes precedence over validation_split.
            It can be:
                - tuple (x_val, y_val) of Numpy arrays or tensors. Preferred data type (less overhead).
                - A tf.data dataset.
    :param do_conf_mat: Generate a confusion matrix based on results.
    :return: (string) Full path to the run_data.csv generated file.
        It can be used by cvnn.data_analysis.SeveralMonteCarloComparison to compare several runs.
    """
    if open_dataset:
        dataset = dp.OpenDataset(open_dataset)  # Warning, open_dataset overwrites dataset

    # Monte Carlo
    monte_carlo = MonteCarlo()
    for model in models:
        # model.training_param_summary()
        monte_carlo.add_model(model)
    if not open_dataset:
        dataset.save_data(monte_carlo.monte_carlo_analyzer.path)
    monte_carlo.run(dataset.x, dataset.y, iterations=iterations,
                    validation_split=validation_split, validation_data=validation_data, do_conf_mat=do_conf_mat,
                    epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                    shuffle=False, debug=debug, data_summary=dataset.summary(), polar=polar)
    if do_all:
        monte_carlo.monte_carlo_analyzer.do_all()

    # Save data to remember later what I did.
    _save_montecarlo_log(iterations=iterations,
                         path=str(monte_carlo.monte_carlo_analyzer.path),
                         models_names=[str(model.name) for model in models],
                         dataset_name=dataset.dataset_name,
                         num_classes=str(dataset.y.shape[1]),
                         polar_mode='Yes' if polar else 'No',
                         dataset_size=str(dataset.x.shape[0]),
                         features_size=str(dataset.x.shape[1]), epochs=epochs, batch_size=batch_size
                         )
    return str(monte_carlo.monte_carlo_analyzer.path / "run_data.csv")


def run_gaussian_dataset_montecarlo(iterations: int = 1000, m: int = 10000, n: int = 128, param_list=None,
                                    epochs: int = 150, batch_size: int = 100, display_freq: int = 1,
                                    optimizer='sgd',       # TODO: Add typing here
                                    shape_raw: List[int] = None, activation: t_activation = 'cart_relu',
                                    debug: bool = False, polar: bool = False, do_all: bool = True,
                                    dropout: Optional[float] = None, models: Optional[List[Model]] = None) -> str:
    """
    This function is used to compare CVNN vs RVNN performance over statistical non-circular data.
        1. Generates a complex-valued gaussian correlated noise with the characteristics given by the inputs.
        2. It then runs a monte carlo simulation of several iterations of both CVNN and an equivalent RVNN model.
        3. Saves several files into ./log/montecarlo/date/of/run/
            3.1. run_summary.txt: Summary of the run models and data
            3.2. run_data.csv: Full information of performance of iteration of each model at each epoch
            3.3. complex_network_statistical_result.csv: Statistical results of all iterations of CVNN per epoch
            3.4. real_network_statistical_result.csv: Statistical results of all iterations of RVNN per epoch
            3.5. (Optional) `plot/` folder with the corresponding plots generated by MonteCarloAnalyzer.do_all()
    :param iterations: Number of iterations to be done for each model
    :param m: Total size of the dataset (number of examples)
    :param n: Number of features / input vector
    :param param_list: A list of len = number of classes.
        Each element of the list is another list of len = 3 with values: [correlation_coeff, sigma_x, sigma_y]
        Example for dataset type A of paper https://arxiv.org/abs/2009.08340:
            param_list = [
                [0.5, 1, 1],
                [-0.5, 1, 1]
            ]
        Default: None will default to the example.
    :param epochs: Number of epochs for each iteration
    :param batch_size: Batch size at each iteration
    :param display_freq: Frequency in terms of epochs of when to do a checkpoint.
    :param optimizer: Optimizer to be used. Keras optimizers are not allowed.
            Can be either cvnn.optimizers.Optimizer or a string listed in opt_dispatcher.
    :param shape_raw: List of sizes of each hidden layer.
        For example [64] will generate a CVNN with one hidden layer of size 64.
        Default None will default to example.
    :param activation: Activation function to be used at each hidden layer
    :param debug:
    :param polar: Boolean weather the RVNN should receive real and imaginary part (False) or amplitude and phase (True)
    :param do_all: If true (default) it creates a `plot/` folder with the plots generated by MonteCarloAnalyzer.do_all()
    :param dropout: (float) Dropout to be used at each hidden layer. If None it will not use any dropout.
    :param models: List of cvnn.CvnnModel to be compared.
    :return: (string) Full path to the run_data.csv generated file.
        It can be used by cvnn.data_analysis.SeveralMonteCarloComparison to compare several runs.
    """
    # Get parameters
    if param_list is None:
        param_list = [
            [0.5, 1, 1],
            [-0.5, 1, 1]
        ]
    dataset = dp.CorrelatedGaussianCoeffCorrel(m, n, param_list, debug=False)
    print("Database loaded...")
    if models is not None:
        return run_montecarlo(models=models, dataset=dataset, open_dataset=None,
                              iterations=iterations, epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                              validation_split=0.2, validation_data=None,
                              debug=debug, polar=polar, do_all=do_all, do_conf_mat=True)
    else:
        return mlp_run_real_comparison_montecarlo(dataset, None, iterations, epochs, batch_size, display_freq,
                                                  optimizer, shape_raw, activation, debug, polar, do_all,
                                                  dropout=dropout)


def mlp_run_real_comparison_montecarlo(dataset: cvnn.dataset.Dataset, open_dataset: Optional[t_path] = None,
                                       iterations: int = 1000,
                                       epochs: int = 150, batch_size: int = 100, display_freq: int = 1,
                                       optimizer='sgd',     # TODO: Typing
                                       shape_raw=None, activation: t_activation = 'cart_relu',
                                       debug:  bool = False, polar: bool = False, do_all: bool = True,
                                       dropout: float = 0.5, validation_split: float = 0.2,
                                       validation_data: Optional[Union[Tuple, data.Dataset]] = None,    # TODO: Add typing of tuple
                                       capacity_equivalent: bool = True, equiv_technique: str = 'ratio',
                                       do_conf_mat: bool = True, checkpoints: bool = False,
                                       shuffle: bool = False) -> str:
    """
    This function is used to compare CVNN vs RVNN performance over any dataset.
    1. Automatically creates two Multi-Layer Perceptrons (MLP), one complex and one real.
    2. Runs simulation and compares them.
    3. Saves several files into ./log/montecarlo/date/of/run/
        3.1. run_summary.txt: Summary of the run models and data
        3.2. run_data.csv: Full information of performance of iteration of each model at each epoch
        3.3. complex_network_statistical_result.csv: Statistical results of all iterations of CVNN per epoch
        3.4. real_network_statistical_result.csv: Statistical results of all iterations of RVNN per epoch
        3.5. (Optional) `plot/` folder with the corresponding plots generated by MonteCarloAnalyzer.do_all()

    :param dataset: cvnn.dataset.Dataset with the dataset to be used on the training
    :param open_dataset: (None)
        If dataset is saved inside a folder and must be opened, path of the Dataset to be opened. Else None (default)
    :param iterations: Number of iterations to be done for each model
    :param epochs: Number of epochs for each iteration
    :param batch_size: Batch size at each iteration
    :param display_freq: Frequency in terms of epochs of when to do a checkpoint.
    :param optimizer: Optimizer to be used. Keras optimizers are not allowed.
            Can be either cvnn.optimizers.Optimizer or a string listed in opt_dispatcher.
    :param shape_raw: List of sizes of each hidden layer.
        For example [64] will generate a CVNN with one hidden layer of size 64.
        Default None will default to example.
    :param activation: Activation function to be used at each hidden layer
    :param debug:
    :param polar: Boolean weather the RVNN should receive real and imaginary part (False) or amplitude and phase (True)
    :param do_all: If true (default) it creates a `plot/` folder with the plots generated by MonteCarloAnalyzer.do_all()
    :param dropout: (float) Dropout to be used at each hidden layer. If None it will not use any dropout.
    :param validation_split: Float between 0 and 1.
            Percentage of the input data to be used as test set (the rest will be use as train set)
            Default: 0.0 (No validation set).
            This input is ignored if validation_data is given.
        :param validation_data: Data on which to evaluate the loss and any model metrics at the end of each epoch.
            The model will not be trained on this data. This parameter takes precedence over validation_split.
            It can be:
                - tuple (x_val, y_val) of Numpy arrays or tensors. Preferred data type (less overhead).
                - A tf.data dataset.
    :param capacity_equivalent: An equivalent model can be equivalent in terms of layer neurons or
                        trainable parameters (capacity equivalent according to: https://arxiv.org/abs/1811.12351)
            - True, it creates a capacity-equivalent model in terms of trainable parameters
            - False, it will double all layer size (except the last one if classifier=True)
    :param equiv_technique: Used to define the strategy of the capacity equivalent model.
        This parameter is ignored if capacity_equivalent=False
        - 'ratio': neurons_real_valued_layer[i] = r * neurons_complex_valued_layer[i], 'r' constant for all 'i'
        - 'alternate': Method described in https://arxiv.org/abs/1811.12351 where one alternates between
                multiplying by 2 or 1. Special case on the middle is treated as a compromise between the two.
    :param do_conf_mat: Generate a confusion matrix based on results.
    :return: (string) Full path to the run_data.csv generated file.
        It can be used by cvnn.data_analysis.SeveralMonteCarloComparison to compare several runs.
    """
    if shape_raw is None:
        shape_raw = [64]
    if open_dataset:
        dataset = dp.OpenDataset(open_dataset)  # Warning, open_dataset overwrites dataset
    # Create complex network
    input_size = dataset.x.shape[1]  # Size of input
    output_size = dataset.y.shape[1]  # Size of output
    shape = [
        layers.ComplexInput(input_shape=input_size)
    ]
    if len(shape_raw) == 0:
        logger.warning("No hidden layers are used. activation and dropout will be ignored")
        shape.append(
            ComplexDense(units=output_size, activation='softmax_real', input_dtype=np.complex64)
        )
    else:  # len(shape_raw) > 0:
        for s in shape_raw:
            shape.append(ComplexDense(units=s, activation=activation))   # Add dropout!
            if dropout is not None:
                shape.append(ComplexDropout(rate=dropout))
        shape.append(ComplexDense(units=output_size, activation='softmax_real'))

    complex_network = tf.keras.Sequential(shape, name="complex_network")
    complex_network.compile(optimizer=optimizer, loss=tf.keras.losses.CategoricalCrossentropy(), metrics=['accuracy'])

    # Monte Carlo
    monte_carlo = RealVsComplex(complex_network,
                                capacity_equivalent=capacity_equivalent, equiv_technique=equiv_technique)
    sleep(1)  # I have error if not because not enough time passed since creation of models to be in diff folders
    monte_carlo.run(dataset.x, dataset.y, iterations=iterations,
                    epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                    shuffle=shuffle, debug=debug, data_summary=dataset.summary(), polar=polar,
                    validation_split=validation_split, validation_data=validation_data, do_conf_mat=do_conf_mat,
                    checkpoints=checkpoints)
    if do_all:
        monte_carlo.monte_carlo_analyzer.do_all()

    # Save data to remember later what I did.
    max_epoch = monte_carlo.pandas_full_data['epoch'].max()
    epoch_filter = monte_carlo.pandas_full_data['epoch'] == max_epoch
    complex_filter = monte_carlo.pandas_full_data['network'] == "complex_network"
    real_filter = monte_carlo.pandas_full_data['network'] == "real_network"
    complex_last_epochs = monte_carlo.pandas_full_data[epoch_filter & complex_filter]
    real_last_epochs = monte_carlo.pandas_full_data[epoch_filter & real_filter]
    complex_median = complex_last_epochs['accuracy'].median()
    real_median = real_last_epochs['accuracy'].median()
    complex_median_train = complex_last_epochs['val_accuracy'].median()
    real_median_train = real_last_epochs['val_accuracy'].median()
    _save_rvnn_vs_cvnn_montecarlo_log(
        iterations=iterations,
        path=str(monte_carlo.monte_carlo_analyzer.path),
        dataset_name=dataset.dataset_name,
        optimizer=str(complex_network.optimizer.__class__),
        loss=str(complex_network.loss.__class__),
        hl=str(len(shape_raw)), shape=str(shape_raw),
        dropout=str(dropout), num_classes=str(dataset.y.shape[1]),
        polar_mode='Yes' if polar else 'No',
        activation=activation,
        dataset_size=str(dataset.x.shape[0]), feature_size=str(dataset.x.shape[1]),
        epochs=epochs, batch_size=batch_size,
        winner='CVNN' if complex_median > real_median else 'RVNN',
        complex_median=complex_median, real_median=real_median,
        complex_median_train=complex_median_train, real_median_train=real_median_train,
        complex_err=median_error(complex_last_epochs['accuracy'].quantile(.75),
                                 complex_last_epochs['accuracy'].quantile(.25), iterations),
        real_err=median_error(real_last_epochs['val_accuracy'].quantile(.75),
                              real_last_epochs['val_accuracy'].quantile(.25), iterations),
        filename='./log/mlp_montecarlo_summary.xlsx'
    )
    return str(monte_carlo.monte_carlo_analyzer.path / "run_data.csv")


# ====================================
#     Excel logging
# ====================================
def _create_excel_file(fieldnames: List[str], row_data: List, filename: Optional[t_path] = None,
                       percentage_cols: Optional[List[str]] = None):
    if filename is None:
        filename = './log/montecarlo_summary.xlsx'
    file_exists = os.path.isfile(filename)
    if file_exists:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
        del ws.tables["Table1"]
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.append(fieldnames)
    ws.append(row_data)
    # TODO: What if len(row_data) is longer than the dictionary? It corresponds with excel's column names?
    tab = Table(displayName="Table1", ref="A1:" + str(chr(64 + len(row_data))) + str(ws.max_row))
    if percentage_cols is not None:
        for col in percentage_cols:
            ws[col + str(ws.max_row)].number_format = '0.00%'
    ws.add_table(tab)
    wb.save(filename)


def _save_rvnn_vs_cvnn_montecarlo_log(iterations, path, dataset_name, hl, shape, dropout, num_classes, polar_mode,
                                      activation, optimizer, loss,
                                      dataset_size, feature_size, epochs, batch_size, winner,
                                      complex_median, real_median, complex_err, real_err,
                                      complex_median_train, real_median_train,
                                      comments='', filename=None):
    fieldnames = ['iterations', 'dataset', '# Classes', "Dataset Size", 'Feature Size', "Polar Mode", "Optimizer",
                  "Loss",
                  'HL', 'Shape', 'Dropout', "Activation Function", 'epochs', 'batch size',
                  "Winner", "CVNN median", "RVNN median", 'CVNN err', 'RVNN err',
                  "CVNN train median", "RVNN train median",
                  'path', "cvnn version", "Comments"
                  ]
    row_data = [iterations, dataset_name, num_classes, dataset_size, feature_size, polar_mode,  # Dataset information
                optimizer, str(loss), hl, shape, dropout, activation, epochs, batch_size,  # Model information
                winner, complex_median, real_median, complex_err, real_err,  # Preliminary results
                complex_median_train, real_median_train,
                path, cvnn.__version__, comments  # Library information
                ]
    percentage_cols = ['P', 'Q', 'R', 'S', 'T', 'U']
    _create_excel_file(fieldnames, row_data, filename, percentage_cols=percentage_cols)


def _save_montecarlo_log(iterations, path, dataset_name, models_names, num_classes, polar_mode, dataset_size,
                         features_size, epochs, batch_size, filename=None):
    fieldnames = [
        'iterations',
        'dataset', '# Classes', "Dataset Size", 'Feature Size',     # Dataset information
        'models', 'epochs', 'batch size', "Polar Mode",             # Models information
        'path', "cvnn version"                                      # Library information
    ]
    row_data = [
        iterations,
        dataset_name, num_classes, dataset_size, features_size,
        '-'.join(models_names), epochs, batch_size, polar_mode,
        path, cvnn.__version__
    ]
    _create_excel_file(fieldnames, row_data, filename)


if __name__ == "__main__":
    # Base case with one hidden layer size 64 and dropout 0.5
    run_gaussian_dataset_montecarlo(iterations=10, dropout=0.5)
